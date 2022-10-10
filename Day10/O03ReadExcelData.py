
from openpyxl import load_workbook
from prettytable import PrettyTable

wb = load_workbook("FirstExcel.xlsx")

wb.active = wb['Players']

ws = wb.active

print(ws.dimensions)

dataRange = ws[ws.dimensions]

row = 1
for c1, c2, c3, c4, c5 in dataRange:
    if row == 1:
        plyrTbl = PrettyTable([c1.value, c2.value, c3.value, c4.value, c5.value])
    else:
        plyrTbl.add_row([c1.value, c2.value, c3.value, c4.value, c5.value])
    row += 1

plyrTbl.align['PlayerName'] = 'l'
plyrTbl.align['Country'] = 'l'

print(plyrTbl)

# print("{0:15}\t\t{1:5}\t\t{2:5}\t\t{3:5}\t\t{4:5}".format(c1.value, c2.value, c3.value, c4.value, c5.value ))
