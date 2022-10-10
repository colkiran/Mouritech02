
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart3D, PieChart3D
from openpyxl.chart.label import DataLabelList

wb = load_workbook("FirstExcel.xlsx")

wb.create_sheet("Chart")

wb.active = wb['Chart']

ws = wb.active

data = [
    ('Products', "Sales"),
    ('Pepsi', 450),
    ('Coke', 380),
    ('Sprite', 300),
    ('Mirinda', 340),
    ('Thumbs up', 485),
    ('Fanta', 270),
    ('Slice', 250)
]

print(f"data :{data}")

for row in data:
    ws.append(row)

dataRef = Reference(ws, min_row=2, min_col=2, max_row=8)
labelRef = Reference(ws, min_row=2, min_col=1, max_row=8)

chart = PieChart3D()
chart.add_data(dataRef)
chart.set_categories(labelRef)
# chart.x_axis.title = "Products"
# chart.y_axis.title = "Salas in lakhs"
chart.title = "Baverage Sales"
chart.dataLabels = DataLabelList()
chart.dataLabels.showVal = True


ws.add_chart(chart, "E9")
wb.save("FirstExcel.xlsx")
