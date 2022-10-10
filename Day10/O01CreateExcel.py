# python => driver (openpyxl)  => excel file

from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

ws = wb.active

ws.title = "Mysheet"

ws["B9"] = "Hello World"
ws["D9"] = 12500
ws["F9"] = datetime.now()
ws["H9"].value = '=column()'

wb.save("FirstExcel.xlsx")